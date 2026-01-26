# -*- coding: utf-8 -*-
"""
Agente de Reportes: Genera DataFrame y Excel con las alertas.
"""

import logging
import tempfile
from datetime import datetime
from pathlib import Path
from typing import Dict, List, Optional, Tuple

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows

import sys
sys.path.insert(0, str(Path(__file__).resolve().parent.parent.parent))

logger = logging.getLogger(__name__)

# Colores para Excel
COLOR_HIGH = "FFCCCC"     # Rojo claro
COLOR_MEDIUM = "FFF4CC"   # Amarillo claro
COLOR_LOW = "CCFFCC"      # Verde claro
COLOR_HEADER = "D9D9D9"   # Gris claro


def create_alerts_dataframe(alerts: List[Dict]) -> pd.DataFrame:
    """
    Crea un DataFrame de pandas con las alertas formateadas.
    
    Args:
        alerts: Lista de alertas detectadas.
    
    Returns:
        pd.DataFrame: DataFrame con las columnas requeridas.
    """
    if not alerts:
        return pd.DataFrame(columns=[
            "Contrato/Expediente",
            "Observación Importante",
            "Acción Requerida",
            "Prioridad"
        ])
    
    rows = []
    for alert in alerts:
        rows.append({
            "Contrato/Expediente": alert.get("expediente", ""),
            "Observación Importante": alert.get("observacion", ""),
            "Acción Requerida": alert.get("accion", ""),
            "Prioridad": alert.get("prioridad", "")
        })
    
    df = pd.DataFrame(rows)
    
    # Ordenar por prioridad (Alta primero)
    priority_order = {"🔴 Alta": 0, "🟡 Media": 1, "🟢 Baja": 2}
    df["_order"] = df["Prioridad"].map(lambda x: priority_order.get(x, 3))
    df = df.sort_values("_order").drop("_order", axis=1)
    df = df.reset_index(drop=True)
    
    return df


def get_priority_color(priority: str) -> str:
    """
    Obtiene el color de fondo según la prioridad.
    
    Args:
        priority: Texto de prioridad con emoji.
    
    Returns:
        str: Código de color hex.
    """
    if "🔴" in priority or "Alta" in priority:
        return COLOR_HIGH
    elif "🟡" in priority or "Media" in priority:
        return COLOR_MEDIUM
    else:
        return COLOR_LOW


def generate_excel_report(
    df: pd.DataFrame,
    alerts_summary: Dict = None,
    output_path: str = None
) -> Tuple[str, bool]:
    """
    Genera un archivo Excel formateado con el informe.
    
    Args:
        df: DataFrame con las alertas.
        alerts_summary: Resumen de alertas (opcional).
        output_path: Ruta de salida (opcional, genera temporal si no se especifica).
    
    Returns:
        Tuple[str, bool]: (ruta del archivo, éxito)
    """
    try:
        # Crear workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Informe_Contratos"
        
        # Estilos
        header_font = Font(bold=True)
        header_fill = PatternFill(start_color=COLOR_HEADER, end_color=COLOR_HEADER, fill_type="solid")
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Escribir headers
        headers = ["Contrato/Expediente", "Observación Importante", "Acción Requerida", "Prioridad"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.border = thin_border
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        
        # Escribir datos
        for row_idx, row in enumerate(df.itertuples(index=False), 2):
            priority = row[3]  # Columna Prioridad
            row_fill = PatternFill(
                start_color=get_priority_color(priority),
                end_color=get_priority_color(priority),
                fill_type="solid"
            )
            
            for col_idx, value in enumerate(row, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.fill = row_fill
                cell.border = thin_border
                cell.alignment = Alignment(vertical='center', wrap_text=True)
        
        # Ajustar anchos de columna
        ws.column_dimensions['A'].width = 25
        ws.column_dimensions['B'].width = 50
        ws.column_dimensions['C'].width = 40
        ws.column_dimensions['D'].width = 15
        
        # Crear hoja de metadatos
        ws_meta = wb.create_sheet("Metadatos")
        meta_data = [
            ("Fecha de generación", datetime.now().strftime("%d/%m/%Y %H:%M:%S")),
            ("Total de contratos procesados", alerts_summary.get("total", 0) if alerts_summary else len(df)),
            ("Alertas Alta Prioridad", alerts_summary.get("high", 0) if alerts_summary else sum(1 for p in df["Prioridad"] if "🔴" in p)),
            ("Alertas Media Prioridad", alerts_summary.get("medium", 0) if alerts_summary else sum(1 for p in df["Prioridad"] if "🟡" in p)),
            ("Alertas Baja Prioridad", alerts_summary.get("low", 0) if alerts_summary else sum(1 for p in df["Prioridad"] if "🟢" in p)),
        ]
        
        for row_idx, (label, value) in enumerate(meta_data, 1):
            ws_meta.cell(row=row_idx, column=1, value=label).font = Font(bold=True)
            ws_meta.cell(row=row_idx, column=2, value=value)
        
        ws_meta.column_dimensions['A'].width = 30
        ws_meta.column_dimensions['B'].width = 25
        
        # Guardar archivo
        if not output_path:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            output_path = str(Path(tempfile.gettempdir()) / f"informe_contratos_{timestamp}.xlsx")
        
        wb.save(output_path)
        logger.info(f"Excel generado: {output_path}")
        
        return output_path, True
        
    except Exception as e:
        logger.error(f"Error generando Excel: {e}")
        return "", False


def run_reporter_node(state: Dict) -> Dict:
    """
    Nodo del grafo LangGraph: Genera el reporte final.
    
    Args:
        state: Estado actual del grafo (debe contener 'alerts').
    
    Returns:
        Dict: Estado actualizado con DataFrame y ruta del Excel.
    """
    logger.info("=== AGENTE REPORTES: Generando informe ===")
    
    alerts = state.get("alerts", [])
    alerts_summary = state.get("alerts_summary", {})
    
    # Crear DataFrame
    df = create_alerts_dataframe(alerts)
    
    # Generar Excel
    excel_path, success = generate_excel_report(df, alerts_summary)
    
    return {
        **state,
        "report_dataframe": df,
        "excel_path": excel_path if success else None,
        "report_complete": True
    }
